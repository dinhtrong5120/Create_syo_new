import warnings

import pandas as pd
import streamlit
from openpyxl import load_workbook
from src.app.list_arrange import check_list_sap_xep

warnings.filterwarnings("ignore")


def find_color_cell(sheet_, column_numbers_):
    gray_cells = []
    blue_cells = []
    delete_cells = []
    blue_check = []
    for column_number in column_numbers_:

        for row_idx in range(1, sheet_.max_row + 1):
            cell = sheet_.cell(row=row_idx, column=column_number)
            #
            if cell.fill.start_color.type != 'theme':
                if cell.value and cell.fill and cell.fill.start_color.index[2:] == '535353':
                    gray_cells.append((cell.value, cell.row))
                    delete_cells.append(cell.row)
                elif cell.fill and cell.fill.start_color.index[2:] == '245269' and cell.value is not None:
                    blue_cells.append((cell.value, cell.row))
                    blue_check.append(cell.value)
                    delete_cells.append(cell.row)
                elif cell.fill and cell.fill.start_color.index[2:] == '969696' and cell.row > 9:
                    delete_cells.append(cell.row)
    blue_check.insert(0, 'UNKNOW_device')
    check_list_sap_xep('./src/db/your_file.txt', blue_check)
    return gray_cells, blue_cells, delete_cells


def find_color_columns(sheet_, column_numbers_, row_idx):
    delete_columns = []
    for column_number in column_numbers_:
        cell = sheet_.cell(row=row_idx, column=column_number)
        if cell.fill.start_color.type != 'theme':
            if cell.fill and cell.fill.start_color.index[2:] == '969696':
                delete_columns.append(cell.column)
    return delete_columns


def unmerge_and_fill(sheet_):
    merged_ranges_and_values = [
        (merged_cell_range, sheet_.cell(merged_cell_range.min_row, merged_cell_range.min_col).value)
        for merged_cell_range in sheet_.merged_cells.ranges
    ]

    df = pd.DataFrame(index=range(1, sheet_.max_row + 1), columns=range(1, sheet_.max_column + 1))
    for row in range(1, sheet_.max_row + 1):
        for col in range(1, sheet_.max_column + 1):
            df.at[row, col] = sheet_.cell(row=row, column=col).value

    for merged_range, value in merged_ranges_and_values:

        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col

        for row in range(min_row, max_row + 1):

            for col in range(min_col, max_col + 1):
                df.at[row, col] = value
    return df


def dataframe_convert(sheet, file_path_, project_name_):
    column_number = [1, 4, 6, 8]
    df = unmerge_and_fill(sheet)

    df.dropna(axis=1, how='all', inplace=True)
    # ************************************************
    result = df.stack()
    index_row_zone = result[result == 'ZONE'].index[0][1]
    column_number_1 = range(index_row_zone, df.shape[1] + 1)
    delete_columns = find_color_columns(sheet, column_number_1, index_row_zone)
    df.drop(columns=delete_columns, inplace=True)

    opt_index = df.shape[0] + 1
    index_column_config_end = df.shape[1]
    gray_cells_list, blue_cells_list_old, delete_cells = find_color_cell(sheet, column_number)
    blue_cells_list_old = [('UNKNOW_device', 0)] + blue_cells_list_old + [('xxx', opt_index)]
    gray_cells_list = [('UNKNOW_device_group', 0)] + gray_cells_list + [('xxx', blue_cells_list_old[-1][1] - 1)]
    delete_cells = list(set(delete_cells))
    delete_cells = sorted(delete_cells)
    set_gray_cells_list_values = {value for _, value in gray_cells_list}
    blue_cells_list = []
    for item in blue_cells_list_old:
        device, value = item
        if (value - 1) in set_gray_cells_list_values:
            blue_cells_list.append((device, value - 1))
        else:
            blue_cells_list.append(item)
    index_gray_old = 0
    value_gray_old = 'xxx'

    for item in gray_cells_list:
        if item[0] == 'UNKNOW_device_group':
            value_gray_old = item[0]
            index_gray_old = item[1]
        else:
            df.loc[index_gray_old:item[1], 'device_group'] = value_gray_old
            value_gray_old = item[0]
            index_gray_old = item[1]
    index_blue_old = 0
    value_blue_old = 'xxx'
    for item in blue_cells_list:
        if item[0] == 'UNKNOW_device':
            value_blue_old = item[0]
            index_blue_old = item[1]
        else:
            df.loc[index_blue_old:item[1], 'device_name'] = value_blue_old
            value_blue_old = item[0]
            index_blue_old = item[1]

    df_end = df.iloc[:opt_index]
    result = df_end.stack()
    index_column_class = result[result == 'Class'].index[0][1]
    index_row_class = result[result == 'Class'].index[0][0]
    index_row_zone = result[result == 'ZONE'].index[0][0]
    index_column_zone = result[result == 'ZONE'].index[0][1]
    index_column_attributes = result[result == 'Attributes'].index[0][1]
    index_column_OptionPackage = result[result == 'Option Package'].index[0][1]
    df_12_begin_row_in_spec = df_end.loc[index_row_zone:index_row_class - 1, index_column_zone:]
    df_12_begin_row_in_spec = df_12_begin_row_in_spec.iloc[:, :-2]
    df_data_in_spec = df_end.iloc[index_row_class:, index_column_class - 1:]
    df_data_in_spec.drop(index=delete_cells, inplace=True)
    num_columns_before = df_12_begin_row_in_spec.shape[1]
    df_12_begin_row_in_spec.dropna(axis=1, how='all', inplace=True)
    num_columns_after = df_12_begin_row_in_spec.shape[1]
    num_deleted_columns = num_columns_before - num_columns_after
    index_column_config_end -= num_deleted_columns
    return df_12_begin_row_in_spec, df_data_in_spec, index_column_class, index_column_zone, index_column_attributes, index_column_OptionPackage, index_column_config_end


def main_process_spec_not_protected(file_path_, project_name_):
    wb = load_workbook(file_path_)
    sheet = wb['1-SPEC']
    df_end_region3_, df_end_1_, index_column_class, index_column_zone, index_column_attributes, index_column_OptionPackage, index_column_config_end = dataframe_convert(
        sheet, file_path_, project_name_)
    return df_end_region3_, [df_end_1_, index_column_class, index_column_zone, index_column_attributes,
                             index_column_OptionPackage, index_column_config_end]

#
# if __name__ == "__main__":
#     file_path = r"C:\Users\KNT21617\Downloads\wz1l_Spec_List.xlsx"
#     project_name = 'wz1l'
#     df_end_region3, df_end_1 = main_process_spec_not_protected(file_path, project_name)
#     df_end_region3.to_excel(r'C:\Users\KNT21617\Downloads\newken\project\data\processed\df_end_region3.xlsx')
#     df_end_1[0].to_excel(r'C:\Users\KNT21617\Downloads\newken\project\data\processed\df_end_1.xlsx')
