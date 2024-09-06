import pandas as pd
import streamlit
from openpyxl import load_workbook
from src.app.check_filed_information import check_syo
import warnings
from src.app.list_arrange import check_list_sap_xep

warnings.filterwarnings("ignore")


def find_color_cell(file_path_, sheet_name, column_number):
    wb = load_workbook(file_path_)
    sheet = wb[sheet_name]
    gray_cells = []
    blue_cells = []
    blue_check = []
    delete_cells = []

    # Loop through rows
    for row_idx in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=column_number)
        if cell.fill.start_color.type != 'theme':
            # Check cell fill color
            if cell.fill and cell.fill.start_color.index[2:] == '535353':
                # Blue color found
                gray_cells.append((cell.value, cell.row - 2))
                delete_cells.append(cell.row - 2)
            elif cell.fill and cell.fill.start_color.index[2:] == '245269' and cell.value is not None:
                blue_cells.append((cell.value, cell.row - 2))
                blue_check.append(cell.value)
                delete_cells.append(cell.row - 2)
    blue_check.insert(0, 'UNKNOW_device')
    # print('blue_check: ', blue_check)
    check_list_sap_xep('./src/db/your_file.txt', blue_check)
    # print('ok')
    return gray_cells, blue_cells, delete_cells


def dataframe_convert(file_path, project_name, result_querry_region_4_gray, result_querry_region_4_blue):
    sheet_name = 'Sheet1'
    column_number = 1
    flg_create_syo = False

    if not isinstance(file_path, pd.DataFrame):
        try:
            df = pd.read_excel(file_path, sheet_name='Sheet1')
        except:
            return f'{project_name}: Sheet name must be "Sheet1"', []
        if 'gr' in df.columns:
            df.rename(columns={'gr': 'Gr', 'keyword': 'Keyword'}, inplace=True)
        list_error, duplicated_elements = check_syo(df)
        if list_error == ['Columns with missing values']:
            return f'{project_name}: Exist Columns with missing values', []
        elif list_error == ['(Gr, Keyword, CADICS ID, AUTO) columns not in data']:
            return f'{project_name}: (Gr, Keyword, CADICS ID, AUTO) columns not in data', []
        else:
            if len(list_error) > 0 and len(duplicated_elements) == 0:
                return f'{project_name}: Form SYO incorrect ', list_error
            elif len(list_error) == 0 and len(duplicated_elements) > 0:
                streamlit.warning(f'{project_name}: CADICS ID duplicated f{duplicated_elements}')
                # return f'{project_name}: CADICS ID duplicated ', duplicated_elements
            elif len(list_error) > 0 and len(duplicated_elements) > 0:
                return f'{project_name}: Form SYO incorrect and CADICS ID duplicated ', [
                    list_error + duplicated_elements]
    else:
        df = file_path.copy()
        flg_create_syo = True
    if df.shape[1] == 6:
        opt_index = df.index[df['CADICS ID'] == 'OptionCode'].values[0]
        gray_cells_list, blue_cells_list_old, delete_cells = find_color_cell(file_path, sheet_name, column_number)
        blue_cells_list_old = [('UNKNOW_device', 0)] + blue_cells_list_old + [('xxx', opt_index)]
        # check_list_sap_xep('../src/db/your_file.txt', blue_cells_list_old)

        gray_cells_list = [('UNKNOW_device_group', 0)] + gray_cells_list + [('xxx', blue_cells_list_old[-1][1] - 1)]

        index_gray_old = 0
        value_gray_old = 'xxx'
        for item in gray_cells_list:
            if item[0] == 'UNKNOW_device_group':
                value_gray_old = item[0]
                index_gray_old = item[1]
            else:
                df.loc[index_gray_old:item[1], 'device_group'] = value_gray_old
                value_gray_old = item[0]
                index_gray_old = item[1]
        index_blue_old = 0
        value_blue_old = 'xxx'
        for item in blue_cells_list_old:
            if item[0] == 'UNKNOW_device':
                value_blue_old = item[0]
                index_blue_old = item[1]
            else:
                df.loc[index_blue_old:item[1], 'device_name'] = value_blue_old
                value_blue_old = item[0]
                index_blue_old = item[1]
        df_end = df.iloc[:opt_index]

        df_cleaned = df_end.dropna(subset=['Gr', 'CADICS ID', 'auto'], how='all')
        df_cleaned.loc[:, 'project_name'] = project_name
        df_cleaned.drop(index=delete_cells, inplace=True)
        return df_cleaned, pd.DataFrame()
    else:
        if flg_create_syo:
            gray_cells_list = []
            blue_cells_list_old = []
            delete_cells = []
            for item in result_querry_region_4_gray:
                index_gray = df.index[df['auto'] == item[0]].tolist()
                if index_gray:
                    gray_cells_list.append((item[0], index_gray[0]))
                    delete_cells.append(index_gray[0])
                else:
                    gray_cells_list.append((item[0], 0))
            gray_cells_list = set(gray_cells_list)
            for item in result_querry_region_4_blue:
                index_blue = df.index[df['auto'] == item[0]].tolist()
                if index_blue:
                    blue_cells_list_old.append((item[0], index_blue[0]))
                    delete_cells.append(index_blue[0])
                else:
                    blue_cells_list_old.append((item[0], 0))
            blue_cells_list_old = set(blue_cells_list_old)
            opt_index = df.index[df['CADICS ID'] == 'OptionCode'].values[0]
        else:
            gray_cells_list, blue_cells_list_old, delete_cells = find_color_cell(file_path, sheet_name, column_number)
            opt_index = df.index[df['CADICS ID'] == 'OptionCode'].values[0]
        blue_cells_list_old = [('UNKNOW_device', 0)] + list(blue_cells_list_old) + [('xxx', opt_index + 2)]
        gray_cells_list = [('UNKNOW_device_group', 0)] + list(gray_cells_list) + [
            ('xxx', blue_cells_list_old[-1][1])]
        blue_cells_list_old = list(set(blue_cells_list_old))
        gray_cells_list = list(set(gray_cells_list))
        set_gray_cells_list_values = {value for _, value in gray_cells_list}
        blue_cells_list = []
        for item in blue_cells_list_old:
            device, value = item
            if item[0] == 'xxx':
                blue_cells_list.append((device, value))
            else:
                if (value - 1) in set_gray_cells_list_values:
                    blue_cells_list.append((device, value))
                else:
                    blue_cells_list.append(item)

        list_columns_name = []
        if 'group_key_map' in df.columns:
            for i, item in enumerate(reversed(df.columns.values)):
                if str(item).startswith('conf-'):
                    column_names = ['comment_' + str(i + 1) for i in range(0, i - 2)] + ['group_key_map', 'default']

                    list_columns_name = list(df.columns.values[:len(df.columns.values) - i]) + column_names
                    break
        else:
            for i, item in enumerate(reversed(df.columns.values)):
                if str(item).startswith('conf-'):
                    column_names = ['comment_' + str(i + 1) for i in range(0, i)]
                    list_columns_name = list(df.columns.values[:len(df.columns.values) - i]) + column_names
                    break
        df.columns = list_columns_name
        index_gray_old = 0
        value_gray_old = 'xxx'

        gray_cells_list = list(set(gray_cells_list))
        gray_cells_list = sorted(gray_cells_list, key=lambda x: x[1])
        blue_cells_list = list(set(blue_cells_list))
        blue_cells_list = sorted(blue_cells_list, key=lambda x: x[1])
        for item in gray_cells_list:
            if item[0] == 'UNKNOW_device_group':
                value_gray_old = item[0]
                index_gray_old = item[1]
            else:
                df.loc[index_gray_old:item[1], 'device_group'] = value_gray_old
                value_gray_old = item[0]
                index_gray_old = item[1]
        index_blue_old = 0
        value_blue_old = 'xxx'
        for item in blue_cells_list:
            if item[0] == 'UNKNOW_device':
                value_blue_old = item[0]
                index_blue_old = item[1]
            else:
                df.loc[index_blue_old:item[1], 'device_name'] = value_blue_old
                value_blue_old = item[0]
                index_blue_old = item[1]
        df_end = df.iloc[:opt_index]
        df_optioncode = df.iloc[opt_index:, ]
        df_end.loc[:, 'project_name'] = project_name.upper()
        if 'gr' in df_end.columns:
            df_cleaned = df_end.dropna(subset=['gr', 'CADICS ID', 'auto'], how='all')
        else:
            df_cleaned = df_end.dropna(subset=['Gr', 'CADICS ID', 'auto'], how='all')
        df_cleaned.drop(index=delete_cells, inplace=True)
        df_cleaned['CADICS ID'] = df_cleaned['CADICS ID'].astype(str).apply(str.upper)
        return df_cleaned, df_optioncode


# 2.----------------------table config--------------------------------
def config_table(df_):
    list_columns_df = df_.columns.values
    conf_list = [item for item in list_columns_df if item.startswith('conf')]
    config_table_df = pd.DataFrame({'config': conf_list})
    print()
    return config_table_df


# 3.---------------------table lot---------------------------------------
def lot_table(df_1_):
    lot_table_ref = pd.DataFrame({'lot_name': df_1_.iloc[1:, df_1_.columns.get_loc('CADICS ID')]})
    lot_table_ref.reset_index(drop=True, inplace=True)
    return lot_table_ref


# 4.-----------------------table comment_column--------------------------------------
def comment_column_table(df_):
    list_columns_df = df_.columns.values
    comment_list = [item for item in list_columns_df if item.startswith('comment')]
    comment_table_df = pd.DataFrame({'Comment': comment_list})
    return comment_table_df


# 5.---------------------------table device--------------------------
def device_table(df_):
    device_table_df = df_[['device_name', 'device_group']]
    df_new = device_table_df.drop_duplicates()
    return df_new


# 6.------------------------table device_detail--------------------------
def device_detail_table(df_):
    if 'Gr' in df_.columns:
        if 'default' not in df_.columns:
            device_detail_df = df_[['device_name', 'CADICS ID', 'Gr', 'Keyword', 'auto']]
        else:
            device_detail_df = df_[['device_name', 'CADICS ID', 'Gr', 'Keyword', 'auto', 'group_key_map', 'default']]
        device_detail_df.rename(
            columns={'CADICS ID': 'device_details_name', 'Gr': "group_detail", 'Keyword': 'option_detail',
                     'auto': 'auto_detail'},
            inplace=True)
    else:
        if 'default' not in df_.columns:
            device_detail_df = df_[['device_name', 'CADICS ID', 'gr', 'keyword', 'auto']]
        else:
            # print(df_.columns)
            device_detail_df = df_[['device_name', 'CADICS ID', 'gr', 'keyword', 'auto', 'group_key_map', 'default']]
        device_detail_df.rename(
            columns={'CADICS ID': 'device_details_name', 'gr': "group_detail", 'keyword': 'option_detail',
                     'auto': 'auto_detail'},
            inplace=True)

    device_detail_df = device_detail_df.dropna(subset=['device_details_name'])
    device_detail_df = device_detail_df[device_detail_df['device_details_name'] != '']
    return device_detail_df.iloc[12:, :]


# 7.---------------------------table information_project------------------------------
def information_project_table(df_):
    if 'Gr' in df_.columns:
        information_project_df = df_.loc[:11, ['CADICS ID', 'Gr', 'Keyword', 'auto']]
    else:
        information_project_df = df_.loc[:11, ['CADICS ID', 'gr', 'keyword', 'auto']]
    return information_project_df


# 8.-------------------table optioncode--------------------------------------
def optioncode_table(df_1_, project_name_):
    conf_columns = [col for col in df_1_.columns if col.startswith('conf-')]
    list_col_to_index = []
    for item in conf_columns:
        list_col_to_index.append(df_1_.columns.get_loc(item))
    optioncode_df = df_1_.iloc[0, list_col_to_index]

    result_df = optioncode_df.reset_index()
    result_df.loc[:, 'project_name'] = project_name_
    result_df.columns = ['config_name', 'optioncode_value', 'project_name']
    return result_df


# 9.-------------------------table project_device-------------------------
def project_device_table(df_, project_name_):
    df_ref = df_[['device_name']]
    project_device = df_ref.drop_duplicates()
    project_device.loc[:, 'project_name'] = project_name_
    return project_device


# 10.--------------------------table value_inf------------------------------------
def value_inf_table(df_):
    list_columns_df = df_.columns.values
    conf_list = [item for item in list_columns_df if item.startswith('conf')]

    value_if_columns = conf_list + ['CADICS ID', 'project_name']
    new_df = df_[value_if_columns]
    df_ref = pd.melt(new_df, id_vars=['CADICS ID', 'project_name'], value_vars=conf_list, var_name='config_name',
                     value_name='Value')
    value_if_df_ref = df_ref[['project_name', 'config_name', 'CADICS ID', 'Value']]
    value_if_df = value_if_df_ref.loc[:, :]
    value_if_df.rename(
        columns={'CADICS ID': 'parameter_name', 'Value': "value"},
        inplace=True)
    return value_if_df


# 11.---------------------------------table status_lot_config---------------------------
def status_lot_config_table(df_1_, project_name_):
    conf_columns = [col for col in df_1_.columns if col.startswith('conf-')]
    all_columns = ['CADICS ID'] + conf_columns
    list_col_to_index = []
    for item in all_columns:
        list_col_to_index.append(df_1_.columns.get_loc(item))
    status_lot_config_df_ref = df_1_.iloc[1:, list_col_to_index]
    status_lot_config_df = pd.melt(status_lot_config_df_ref, id_vars=['CADICS ID'], value_vars=conf_columns,
                                   var_name='config_name', value_name='Value')
    status_lot_config_df.loc[:, 'project_name'] = project_name_
    status_lot_config_df.rename(columns={'CADICS ID': 'lot_name', 'Value': "status"}, inplace=True)
    return status_lot_config_df


# 12.---------------------------------table project_device_comment--------------------
def project_device_comment_table(df_, project_name_):
    comment_columns = [col for col in df_.columns if col.startswith('comment')]
    if len(comment_columns) > 0:
        all_columns = ['device_name', 'CADICS ID'] + comment_columns
        list_col_to_index = []
        for item in all_columns:
            list_col_to_index.append(df_.columns.get_loc(item))
        project_device_comment_ref = df_.iloc[1:, list_col_to_index]
        print('project_device_comment_ref columns: ', project_device_comment_ref.columns)
        project_device_comment_df = pd.melt(project_device_comment_ref, id_vars=['device_name', 'CADICS ID'],
                                            value_vars=comment_columns,
                                            var_name='Comment', value_name='Value')
        print('project_device_comment_df columns: ', project_device_comment_df.columns)
        project_device_comment_df.loc[:, 'project_name'] = project_name_
        project_device_comment_df.rename(
            columns={'CADICS ID': 'device_details_name', 'Comment': "comment_name", 'Value': 'comment_detail'},
            inplace=True)
        project_device_comment_df.dropna(subset=['device_details_name'], inplace=True)
        project_device_comment_df = project_device_comment_df[project_device_comment_df['device_details_name'] != '']
        return project_device_comment_df
    else:
        return pd.DataFrame()


# 13.-------------------------table status_config_device_detail------------------------------
def status_config_device_detail_table(df_, project_name_):
    config_columns = [col for col in df_.columns if col.startswith('conf-')]
    all_columns = ['device_name', 'CADICS ID'] + config_columns
    list_col_to_index = []
    for item in all_columns:
        list_col_to_index.append(df_.columns.get_loc(item))
    status_config_device_detail_ref = df_.iloc[:, list_col_to_index]
    status_config_device_detail_df = pd.melt(status_config_device_detail_ref, id_vars=['device_name', 'CADICS ID'],
                                             value_vars=config_columns,
                                             var_name='Config', value_name='Value')
    status_config_device_detail_df.loc[:, 'project_name'] = project_name_
    status_config_device_detail_df.rename(
        columns={'CADICS ID': 'device_details_name', 'Config': "config_name", 'Value': 'status'},
        inplace=True)
    status_config_device_detail_df.dropna(subset=['device_details_name'], inplace=True)
    status_config_device_detail_df = status_config_device_detail_df[
        status_config_device_detail_df['device_details_name'] != '']
    return status_config_device_detail_df
